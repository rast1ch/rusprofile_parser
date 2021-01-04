import requests
import openpyxl
from bs4 import BeautifulSoup
import lxml
import TOCKENS


def take_away():
    r = 0
    c = 1
    xyxl = openpyxl.load_workbook('pyth.xlsx',)
    xyxl = xyxl.active
    for i in TOCKENS.URL_MOD:
        for z in range(50000):
            url = TOCKENS.URL_STAT + '/ip/' + i + str(z) 
            r = requests.get(url, headers = TOCKENS.HEADER)
            soup = BeautifulSoup(r.content,'lxml')
            for i in soup.find_all('div', {'class' :'company-item'}):
                if i.find('span', {'class' : 'warning-text'}):
                    break
                else:
                    for x in i.find_all('a'):
                        xyxl.cell(r + 1, c ).value = x.text().strip()
                    r = 0
                    for z in i.find_all('dd'):
                        if (r + 1) % 2 == 1:
                            xyxl.cell(r + 1 , 2).value = z.text().strip()
                        else:
                            xyxl.cell(r + 1 , 3).value = z.text().strip()
                    
                        
            

        
        
if __name__ == "__main__":
    while True:
        print(take_away())
        print('___________________________________________________')